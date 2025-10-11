from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from io import BytesIO
from datetime import datetime
from typing import List, Dict
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def crear_pdf_proyectos(proyectos: List[Dict]) -> BytesIO:
    """Genera un PDF con la lista de proyectos"""
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5*inch, bottomMargin=0.5*inch)
    elements = []
    
    # Estilos
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#1e40af'),
        spaceAfter=12,
        alignment=TA_CENTER
    )
    
    # Título
    titulo = Paragraph("REPORTE DE PROYECTOS DE GRADO", title_style)
    elements.append(titulo)
    elements.append(Spacer(1, 0.2*inch))
    
    # Fecha de generación
    fecha_style = ParagraphStyle('fecha', parent=styles['Normal'], fontSize=10, alignment=TA_RIGHT)
    fecha = Paragraph(f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}", fecha_style)
    elements.append(fecha)
    elements.append(Spacer(1, 0.3*inch))
    
    # Datos de la tabla
    data = [['Código', 'Nombre', 'Estado', 'Presupuesto\nEstimado', 'Presupuesto\nAsignado', 'Presupuesto\nEjecutado']]
    
    for p in proyectos:
        data.append([
            str(p.get('codigo_proyecto', '')),
            str(p.get('nombre', ''))[:30],
            str(p.get('estado', '')),
            f"${float(p.get('presupuesto_estimado', 0)):,.2f}",
            f"${float(p.get('presupuesto_asignado', 0) or 0):,.2f}",
            f"${float(p.get('presupuesto_ejecutado', 0)):,.2f}"
        ])
    
    # Crear tabla
    tabla = Table(data, colWidths=[1.2*inch, 2*inch, 1*inch, 1.2*inch, 1.2*inch, 1.2*inch])
    tabla.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e40af')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
    ]))
    
    elements.append(tabla)
    
    # Total de proyectos
    elements.append(Spacer(1, 0.3*inch))
    total_style = ParagraphStyle('total', parent=styles['Normal'], fontSize=12, textColor=colors.HexColor('#1e40af'))
    total = Paragraph(f"<b>Total de proyectos: {len(proyectos)}</b>", total_style)
    elements.append(total)
    
    doc.build(elements)
    buffer.seek(0)
    return buffer

def crear_excel_proyectos(proyectos: List[Dict]) -> BytesIO:
    """Genera un archivo Excel con la lista de proyectos"""
    buffer = BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Proyectos"
    
    # Estilos
    header_fill = PatternFill(start_color="1e40af", end_color="1e40af", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Encabezados
    headers = ['Código', 'Nombre', 'Estado', 'Descripción', 'Objetivos', 
               'Presupuesto Estimado', 'Presupuesto Asignado', 'Presupuesto Ejecutado', 
               'Fecha Creación', 'Última Modificación']
    
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
    
    # Datos
    for row_idx, p in enumerate(proyectos, start=2):
        ws.cell(row=row_idx, column=1, value=p.get('codigo_proyecto', '')).border = border
        ws.cell(row=row_idx, column=2, value=p.get('nombre', '')).border = border
        ws.cell(row=row_idx, column=3, value=p.get('estado', '')).border = border
        ws.cell(row=row_idx, column=4, value=p.get('descripcion', '')).border = border
        ws.cell(row=row_idx, column=5, value=p.get('objetivos', '')).border = border
        ws.cell(row=row_idx, column=6, value=float(p.get('presupuesto_estimado', 0))).border = border
        ws.cell(row=row_idx, column=6, value=float(p.get('presupuesto_estimado', 0))).number_format = '"$"#,##0.00'
        ws.cell(row=row_idx, column=7, value=float(p.get('presupuesto_asignado', 0) or 0)).border = border
        ws.cell(row=row_idx, column=7).number_format = '"$"#,##0.00'
        ws.cell(row=row_idx, column=8, value=float(p.get('presupuesto_ejecutado', 0))).border = border
        ws.cell(row=row_idx, column=8).number_format = '"$"#,##0.00'
        
        fecha_creacion = p.get('fecha_creacion')
        if fecha_creacion:
            if isinstance(fecha_creacion, str):
                ws.cell(row=row_idx, column=9, value=fecha_creacion).border = border
            else:
                ws.cell(row=row_idx, column=9, value=fecha_creacion.strftime('%d/%m/%Y %H:%M')).border = border
        
        fecha_modificacion = p.get('fecha_ultima_modificacion')
        if fecha_modificacion:
            if isinstance(fecha_modificacion, str):
                ws.cell(row=row_idx, column=10, value=fecha_modificacion).border = border
            else:
                ws.cell(row=row_idx, column=10, value=fecha_modificacion.strftime('%d/%m/%Y %H:%M')).border = border
    
    # Ajustar anchos de columna
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 50
    ws.column_dimensions['E'].width = 50
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 18
    ws.column_dimensions['H'].width = 18
    ws.column_dimensions['I'].width = 20
    ws.column_dimensions['J'].width = 20
    
    wb.save(buffer)
    buffer.seek(0)
    return buffer

def crear_pdf_gastos(gastos: List[Dict]) -> BytesIO:
    """Genera un PDF con la lista de gastos"""
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5*inch, bottomMargin=0.5*inch)
    elements = []
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#dc2626'),
        spaceAfter=12,
        alignment=TA_CENTER
    )
    
    titulo = Paragraph("REPORTE DE GASTOS", title_style)
    elements.append(titulo)
    elements.append(Spacer(1, 0.2*inch))
    
    fecha_style = ParagraphStyle('fecha', parent=styles['Normal'], fontSize=10, alignment=TA_RIGHT)
    fecha = Paragraph(f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}", fecha_style)
    elements.append(fecha)
    elements.append(Spacer(1, 0.3*inch))
    
    data = [['ID', 'Proyecto', 'Concepto', 'Monto', 'Estado', 'Fecha']]
    
    for g in gastos:
        data.append([
            str(g.get('id', '')),
            str(g.get('proyecto_id', '')),
            str(g.get('concepto', ''))[:30],
            f"${float(g.get('monto', 0)):,.2f}",
            str(g.get('estado', '')),
            g.get('creado_en', '')[:10] if isinstance(g.get('creado_en'), str) else g.get('creado_en', datetime.now()).strftime('%d/%m/%Y')
        ])
    
    tabla = Table(data, colWidths=[0.6*inch, 0.8*inch, 2.5*inch, 1.2*inch, 1*inch, 1.2*inch])
    tabla.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#dc2626')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
    ]))
    
    elements.append(tabla)
    elements.append(Spacer(1, 0.3*inch))
    
    total_monto = sum(float(g.get('monto', 0)) for g in gastos)
    total_style = ParagraphStyle('total', parent=styles['Normal'], fontSize=12, textColor=colors.HexColor('#dc2626'))
    total = Paragraph(f"<b>Total de gastos: {len(gastos)} | Monto total: ${total_monto:,.2f}</b>", total_style)
    elements.append(total)
    
    doc.build(elements)
    buffer.seek(0)
    return buffer

def crear_excel_gastos(gastos: List[Dict]) -> BytesIO:
    """Genera un archivo Excel con la lista de gastos"""
    buffer = BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Gastos"
    
    header_fill = PatternFill(start_color="dc2626", end_color="dc2626", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    headers = ['ID', 'Proyecto ID', 'Estudiante ID', 'Concepto', 'Monto', 
               'Estado', 'Soporte URL', 'Comentarios', 'Fecha Creación', 'Fecha Aprobación']
    
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    for row_idx, g in enumerate(gastos, start=2):
        ws.cell(row=row_idx, column=1, value=g.get('id', '')).border = border
        ws.cell(row=row_idx, column=2, value=g.get('proyecto_id', '')).border = border
        ws.cell(row=row_idx, column=3, value=g.get('estudiante_id', '')).border = border
        ws.cell(row=row_idx, column=4, value=g.get('concepto', '')).border = border
        ws.cell(row=row_idx, column=5, value=float(g.get('monto', 0))).border = border
        ws.cell(row=row_idx, column=5).number_format = '"$"#,##0.00'
        ws.cell(row=row_idx, column=6, value=g.get('estado', '')).border = border
        ws.cell(row=row_idx, column=7, value=g.get('soporte_url', '')).border = border
        ws.cell(row=row_idx, column=8, value=g.get('comentarios', '')).border = border
        
        creado_en = g.get('creado_en')
        if creado_en:
            if isinstance(creado_en, str):
                ws.cell(row=row_idx, column=9, value=creado_en).border = border
            else:
                ws.cell(row=row_idx, column=9, value=creado_en.strftime('%d/%m/%Y %H:%M')).border = border
        
        aprobado_en = g.get('aprobado_en')
        if aprobado_en:
            if isinstance(aprobado_en, str):
                ws.cell(row=row_idx, column=10, value=aprobado_en).border = border
            else:
                ws.cell(row=row_idx, column=10, value=aprobado_en.strftime('%d/%m/%Y %H:%M')).border = border
    
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 40
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 40
    ws.column_dimensions['H'].width = 30
    ws.column_dimensions['I'].width = 20
    ws.column_dimensions['J'].width = 20
    
    wb.save(buffer)
    buffer.seek(0)
    return buffer
